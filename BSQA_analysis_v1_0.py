# -*- coding: utf-8 -*-
"""
Created on Wed Mar 30 09:54:51 2022

@author: James.Harkin
"""

import configparser
import pydicom
from pydicom.tag import Tag
from pydicom.filereader import InvalidDicomError
import os
import shutil
import tkinter as tk
import numpy as np
from matplotlib import pyplot as plt
from PIL import ImageTk,Image
from tkinter import ttk, filedialog
import skimage.segmentation as seg
import openpyxl
import cv2      
from fpdf import FPDF 
import webbrowser   
                

root = tk.Tk()
root.title("Root window")
root.geometry("800x800")

class main: 
    #config file to initialise analysis should be in the same folder as this script.
    config_file_name = "BSQA_config.ini"
    class initialise_analysis:
        def __init__(self, default_paths, sequence_names_dict, n_elements):
            """
            Parameters
            ----------
            default_paths : Dictionary of file default paths
            sequence_names_dict : Dictionary of sequence names - only Dicoms with a sequence name in this dictionary will be sorted.
            n_elements: Int  - Number of elements in the coil
            
            """
            
            #ask the user what analysis they want to perform
            self.Setup_Analysis(default_paths["default_archive_path"], default_paths["default_results_path"], default_paths["default_report_path"])
            #ask the user for the root directory containg all DICOMS of interest
            self.images_directory = self.Select_Directory(default_paths["default_search_path"], 
                "Select directory containing all dicom images to be analysed.")
            
            #modify sequence names_dict so only the images required are put in dcm_dict
            if self.SNR_analysis != True:
                del sequence_names_dict["T2"]
            if self.sup_analysis != True:
                del sequence_names_dict["T1"]
                del sequence_names_dict["T1_sup"]
            
            #get all file paths in the selected directory
            self.file_paths = self.Get_Files(self.images_directory)
            #place all dicoms with a valid sequence name in a dictionary
            dcm_dict = self.Get_Dicom_Dict(self.file_paths, sequence_names_dict)
            #sort the dictionary
            self.sorted_dcm_dict = self.Sort_Dicoms(dcm_dict, n_elements)
  
            
            
            if self.archive_images == True:
                # archive sorted images
                if self.user_defined_archive_path == True:
                    self.archive_path = self.Select_Directory(default_paths["default_archive_path"], 
                    "Select directory to archive DICOM images to.")
                else:
                    self.archive_path = default_paths["default_archive_path"]
                self.Export_Images_To_Archive(self.archive_path, self.sorted_dcm_dict, self.acquisition_dates[0])
            
            if self.export_results == True:
                if self.user_defined_results_path == True:
                    #ask user what spreadsheet they want to export to
                    self.results_path = self.Select_File(os.path.split(default_paths["default_results_path"])[0], 
                    "Select spreadsheet to export results to.")
                else:
                    self.results_path = default_paths["default_results_path"]
            else:
                #still need to define results path to produce graphs
                self.results_path = default_paths["default_results_path"]
                
            #setup path to produce report in 
            if self.produce_report == True:
                if self.user_defined_report_path == True:
                    self.report_path = self.Select_Directory(default_paths["default_report_path"], 
                    "Select directory to export report to.")
                else:
                    self.report_path = default_paths["default_report_path"]
            
                self.baseline_path = default_paths["default_baseline_path"]
        
        def Setup_Analysis(self, current_archive_path, current_results_path, current_report_path):
            #ask the user what analysis they want to perform, if they want to archive the images, if they want to export the results and if they want to produce a pdf" 
            
            start_analysis = tk.IntVar()               
            #break out when tk button "Perform Analysis" is selected
            while start_analysis.get() != True:
                top = tk.Toplevel()
                top.title("Confirm Analysis To Perform")
                #top.geometry('300x150')
                top.wm_attributes('-topmost', 1)
                
                
                #set up variables with default values
                SNR_analysis = tk.IntVar()
                SNR_analysis.set(True)
                sup_analysis = tk.IntVar()
                sup_analysis.set(True)
                
                archive_images = tk.IntVar()
                archive_images.set(True)
                user_defined_archive_path = tk.IntVar()
                
                export_results = tk.IntVar()
                export_results.set(True)
                user_defined_results_path = tk.IntVar()
                
                produce_report = tk.IntVar()
                produce_report.set(True)
                user_defined_report_path = tk.IntVar()
                
                
                tk.Label(top, text="Analysis to perform",font=("Courier", 25)).grid(columnspan=2, row=0, column=0)
                tk.Checkbutton(top, text="SNR", variable=SNR_analysis, font=("Courier", 15)).grid(row=1, column=0)
                tk.Checkbutton(top, text="Suppression Effectiveness", variable=sup_analysis, font=("Courier", 15) ).grid(row=1, column=1)
                
                ttk.Separator(top, orient="horizontal").grid(columnspan=2,row=2, sticky="ew")
                
                tk.Label(top, text="Archive Parameters",font=("Courier", 25)).grid(columnspan=2, row=3, column=0)
                tk.Label(top, text="Current archive path: " + current_archive_path, font=("Courier",10)).grid(columnspan=2,row=4, column=0)
                tk.Checkbutton(top, text="Archive images?", variable=archive_images, font=("Courier", 15)).grid(row=5, column=0)
                tk.Checkbutton(top, text="Modify archive path?", variable=user_defined_archive_path, font=("Courier", 15)).grid(row=5, column=1)
                
                ttk.Separator(top, orient="horizontal").grid(columnspan=2,row=6, sticky="ew")
                
                tk.Label(top, text="Result Export Parameters",font=("Courier", 25)).grid(columnspan=2, row=7, column=0)                
                tk.Label(top, text="Current result export path: " + current_results_path, font=("Courier",10)).grid(columnspan=2,row=8, column=0)
                tk.Checkbutton(top, text="Export Results?", variable=export_results, font=("Courier", 15)).grid(row=9, column=0)
                tk.Checkbutton(top, text="Modify result export path?", variable=user_defined_results_path, font=("Courier", 15)).grid(row=9, column=1)
                
                ttk.Separator(top, orient="horizontal").grid(columnspan=2,row=10, sticky="ew")
                
                tk.Label(top, text="Report Export Parameters",font=("Courier", 25)).grid(columnspan=2, row=11, column=0)                
                tk.Label(top, text="Current report export path: " + current_report_path, font=("Courier",10)).grid(columnspan=2,row=12, column=0)
                tk.Checkbutton(top, text="Produce Report", variable=produce_report, font=("Courier", 15)).grid(row=13, column=0)
                tk.Checkbutton(top, text="Modify report export path?", variable=user_defined_report_path, font=("Courier", 15)).grid(row=13, column=1)
                
                ttk.Separator(top, orient="horizontal").grid(columnspan=2,row=14, sticky="ew")
                
                
                
                tk.Button(top, text="Perform Analysis", font=("Courier",15), command=lambda:  start_analysis.set(True)).grid(columnspan=2,row=15, column=0)
                
                #wait until perform analysis button is pressed
                top.wait_variable(start_analysis)                   
                top.destroy()
            
            #once user has selected the analysis parameters define class variables    
            self.SNR_analysis = SNR_analysis.get()
            self.sup_analysis = sup_analysis.get()
            
            self.archive_images = archive_images.get()
            self.user_defined_archive_path = user_defined_archive_path.get()
            
            self.export_results = export_results.get()
            self.user_defined_results_path = user_defined_results_path.get()
            
            self.produce_report = True
            self.user_defined_report_path = user_defined_report_path.get()
            
        def Select_Directory(self, initial_dir, window_title):
            """
            Parameters
            ----------
            initial_dir: str
                Path at which the file dialogue opens
            window_title: str
                Title of the filedialogue window.
                
            Returns
            -------
            user_selected_directory : str
            
            Uses tkinter filedialog.  Asks the user to select a directory and
            returns the user_selected_directory.
            """
            directory_window = tk.Toplevel()
            directory_window.wm_attributes('-topmost', 1)
            directory_window.withdraw()
            user_selected_directory =  tk.filedialog.askdirectory(parent=directory_window, 
                                                               initialdir=initial_dir, 
                                                               title=window_title)
            directory_window.destroy()
            return user_selected_directory
        
        def Select_File(self, initial_dir, window_title):
            """
            Parameters
            ----------
            initial_dir: str
                Path at which the file dialogue opens
            window_title: str
                Title of the filedialogue window.
                
            Returns
            -------
            user_selected_directory : str
            
            Uses tkinter filedialog.  Asks the user to select a directory and
            returns the user_selected_directory.
            """
            file_window = tk.Toplevel()
            file_window.wm_attributes('-topmost', 1)
            file_window.withdraw()
            user_selected_path =  tk.filedialog.askopenfilename(parent=file_window, 
                                                               initialdir=initial_dir, 
                                                               title=window_title)
            file_window.destroy()
            return user_selected_path
        
        def Get_Files(self, starting_dir):
            """
            Parameters
            ----------
            starting_dir: str
                Every file not called VERSION inside 
                starting_dir will be assumed to be a dicom of interest.
            
            Returns
            -------
            paths : list of str
                list of dicom_paths inside starting_dir.
            """
            paths = []
             
            for dirName, subdirList, fileList in os.walk(starting_dir):
                #individual channel data
                for filename in fileList:
                    paths.append(os.path.join(dirName,filename)) 
            return paths
        
        def Get_Dicom_Dict(self, dicom_paths, sequence_names_dict):
            """
            Parameters
            ----------
            dicom_paths : list of strings
                List of all dicom paths
            Returns
            -------
            dcm_dict : Dictionary of dicoms of the form {<sequence type>:{<path>:<DICOM>}}
            """
            #invert sequence names_dict
            sequence_type_dict = {}
            dcm_dict = {}
            for sequence_type in sequence_names_dict: 
                sequence_type_dict[sequence_names_dict[sequence_type]] = sequence_type
            
            for dicom_path in dicom_paths:
                try:
                    dcm = pydicom.read_file(dicom_path)
                    try:
                        dcm_type = sequence_type_dict[dcm.SeriesDescription]
                        try:
                            dcm_dict[dcm_type][dicom_path] = dcm
                        except KeyError:
                            #first DICOM of that type
                            dcm_dict[dcm_type] = {}
                            dcm_dict[dcm_type][dicom_path] = dcm
                    except KeyError:
                        #not a sequence of interest
                        pass
                except InvalidDicomError:
                    #file is not a dicom
                    pass
            return dcm_dict
        
        def Sort_I_E_DICOMs(self, dcm_dict, n_elements):
            """
            Parameters
            ----------
            dcm_dict : Dictionary of dicoms of the form {<sequence type>:{<path>:<DICOM>}}
            n_elements : number of elements in the coil  (if individual element 
                      images haven't been generated specify 1 for a combined element image')
            
            Returns
            -------
            sorted_dcm_dict : Dictionary of dicoms of the form 
            {<series_time>:{<repeat_n>:{"CE":{"pixel_array":<combined element pixel array},
                                            "IE":{<element_n>:{"path":<DICOM Path>, "dcm":<DICOM>}}}}}
            
            
            Sort Individual element dicom images and produce combined element pixel array usind SoS method
            If n_elements=1 than the dcm is assumed to be a combined element 
            image and the sorted dcm dict is only populated with CE pixel arrays
            
            
            """
            
            
            sorted_dcm_dict = {}
            for path in dcm_dict:
                dcm = dcm_dict[path]
                try:
                    if dcm.SeriesDate not in self.acquisition_dates:
                        #imges acquired on multiple dates - will be treated as one dataset if you continue THIS ERROR NEEDS HANDLING
                        print("Images acquired on multiple dates - will be treated as one dataset if you continue.")
                        self.acquisition_dates.append(dcm.SeriesDate)
                except AttributeError:
                    #acquisition Dates not initialised
                    self.acquisition_dates = [dcm.SeriesDate]
                
                #DICOM tag used to differentiate between elements    
                
                #t1 = Tag(0x7a1103e)
                t1 = Tag(0x0200013)
                acquisition_number = dcm[t1].value
                if dcm.SeriesTime not in sorted_dcm_dict.keys():
                    sorted_dcm_dict[dcm.SeriesTime] = {}
                    sorted_dcm_dict[dcm.SeriesTime]["unsorted"] = {}
                #seperate out the series. Each series is assumed to be a dynamic acquisition with 2 images for each element(2 repeats)
                sorted_dcm_dict[dcm.SeriesTime]["unsorted"][acquisition_number] = {"path":path, "dcm":dcm}
                
            for series_time in sorted_dcm_dict:
                calculated_elements = int(len(sorted_dcm_dict[series_time]["unsorted"])/(2))
                if calculated_elements != n_elements:
                    raise Exception("Incorrect number of elements: " + str(calculated_elements) + " calculated")
                               
                SoS_elements = {}
                
                if n_elements != 1:
                    for image_n in sorted_dcm_dict[series_time]["unsorted"]: 
                        element_n = int(image_n % n_elements)+1
                        #should be 2 repeats
                        repeat_n = int((image_n-1)/n_elements)
                        if repeat_n not in sorted_dcm_dict[series_time].keys():
                            sorted_dcm_dict[series_time][repeat_n] = {}
                            sorted_dcm_dict[series_time][repeat_n]["IE"] = {}
                            sorted_dcm_dict[series_time][repeat_n]["CE"] = {}
                        sorted_dcm_dict[series_time][repeat_n]["IE"][element_n] = sorted_dcm_dict[series_time]["unsorted"][image_n]   
                        
                        dcm = sorted_dcm_dict[series_time]["unsorted"][image_n]["dcm"]
                        
                        #create a sum of squares pixel array of all the individual element images 
                        try:
                            SoS_elements[repeat_n] = np.add(SoS_elements[repeat_n],np.square(dcm.pixel_array.astype(float)))
                        except KeyError:
                            SoS_elements[repeat_n] = np.square(dcm.pixel_array.astype(float))
                                    
                        
                    for repeat_n in SoS_elements:
                        sorted_dcm_dict[series_time][repeat_n]["CE"]["pixel_array"] = np.sqrt(SoS_elements[repeat_n])
                    del SoS_elements  
                else:
                    #assumed to be a combined element image
                    for image_n in sorted_dcm_dict[series_time]["unsorted"]:
                        repeat_n = image_n - 1
                        sorted_dcm_dict[series_time][repeat_n] = {}
                        sorted_dcm_dict[series_time][repeat_n]["CE"] = {}
                        sorted_dcm_dict[series_time][repeat_n]["CE"] = sorted_dcm_dict[series_time]["unsorted"][image_n]
                        #create pixel array
                        sorted_dcm_dict[series_time][repeat_n]["CE"]["pixel_array"] = sorted_dcm_dict[series_time]["unsorted"][image_n]["dcm"].pixel_array.astype(float)
                        
                del sorted_dcm_dict[series_time]["unsorted"]
                       
            return sorted_dcm_dict
        
        def Sort_Dicoms(self, dcm_dict, n_elements):
            """
            Parameters
            ----------
            dcm_dict : Dictionary of dicoms of the form {<sequence type>:{<path>:<DICOM>}}
            n_elements : number of elements in the coil  (if individual element 
                      images haven't been generated specify 1 for a combined element image')
            
            Returns
            -------
            sorted_dcm_dict_full : Dictionary of the form {sequence_type:sorted_dcm_dict}

            """
            #sort dicoms for each sequence type into one dictionary sorted_dcm_dict_full
            sorted_dcm_dict_full = {}
            for sequence_type in dcm_dict:
                if sequence_type == "T2":
                    #sort SNR files
                    sorted_dcm_dict_full["T2"] = self.Sort_I_E_DICOMs(dcm_dict["T2"], n_elements)
                elif sequence_type == "T1":
                    #sort T1 nonsupressed files
                    sorted_dcm_dict_full["T1"] = self.Sort_I_E_DICOMs(dcm_dict["T1"], 1)
                    pass
                elif sequence_type == "T1_sup":
                    #sort T1 supressed files
                    sorted_dcm_dict_full["T1_sup"] = self.Sort_I_E_DICOMs(dcm_dict["T1_sup"], 1)
                    pass
            return sorted_dcm_dict_full
    
        def Initialise_Directory(self, path):
            """
            Checks if the path exists. If it doesn't create it.
            
            Parameters
            ----------
            path : str
                Should be a directory
            """
            if not os.path.exists(path):
                os.makedirs(path)
        
        def Copy_File(self, current_path, target_directory, new_name):
            """
            Checks if a file already exists in the new location.  If it 
            doesn't the file is copied to the new location and given a new 
            name.
            
            Parameters
            ----------
            current_path : str
                Current location of file
            target_directory : str
                Directory to copy the file too.
            new_name : str
                New name of the file

            """
            self.new_path = os.path.join(target_directory, new_name)                                    
            if not os.path.exists(self.new_path):
                self.Initialise_Directory(target_directory)
                shutil.copyfile(current_path, self.new_path)
    
        def Export_Images_To_Archive(self, archive_root, dcm_dict, date):
            """export images in dcm_dict to archive.  Files are exported to the folder:
                archive_root\<date>\<sequence_type>\<series_ID>\
            """
            
            for sequence_type in dcm_dict:
                for series_time in dcm_dict[sequence_type]:
                    for repeat_n in dcm_dict[sequence_type][series_time]:
                        try:
                            #try exporting the IE dcms - if they dont exist export "CE" dcms
                            for array_n in dcm_dict[sequence_type][series_time][repeat_n]["IE"]:
                                current_path = dcm_dict[sequence_type][series_time][repeat_n]["IE"][array_n]["path"]
                                series_ID = str(series_time) + "_" + str(repeat_n)
                                target_directory =  os.path.join(archive_root, date, sequence_type, series_ID)
                                new_name = "Element_" + str(array_n)
                                self.Copy_File(current_path, target_directory, new_name)
                        except KeyError:
                            #only entered if IE dcms dont exist
                            current_path = dcm_dict[sequence_type][series_time][repeat_n]["CE"]["path"]
                            series_ID = str(series_time) + "_" + str(repeat_n)
                            target_directory =  os.path.join(archive_root, date, sequence_type, series_ID)
                            new_name = "Combined_Elements"
                            self.Copy_File(current_path, target_directory, new_name)
                         
    class initialise_masks:
        """
        Assumption made that all images selected are in the same location
        Masks are initially produced using the thresholds established in the class initialise_analysis.
        The visualised mask should be created using the combined element image
        and the ideal mask should cover the entire phantom.
        The same mask is used for the T1w and T2w images.
        To produce the phantom mask the accepted mask is contracted by 4 pixels and to 
        produce the in air mask the accepted mask is inverted and contracted by 5 pixels.
        
        initialises self.mask_dict  (of the form {<mask type>:<mask>})
        TM_1_mask should cover the not suppressed tissue mimick
        TM_2_mask should cover the suppressed tissue mimick
        
        """
        #low pass filter used to smooth  out noise in image
        low_pass_filter = np.array([[1,2,1],[2,4,2],[1,2,1]])/16
        def __init__(self, dcm_dict, coil_spec):
            self.mask_dict = {}
            self.lower_threshold = float(coil_spec["Lower_Threshold"])
            self.upper_threshold = float(coil_spec["Upper_Threshold"])
            
            #combined image required for generating mask: loop through dictionary to find one
            #SNR maskcalculated over entire phantom
            #try creating masks using T2w image - if not use T1w image
            try:
                for series_time in dcm_dict["T2"]:
                    for acq_time in dcm_dict["T2"][series_time]:
                        #image to establish mask from
                        dcm_array = dcm_dict["T2"][series_time][acq_time]["CE"]["pixel_array"].copy()
                        #image mask is the mask visualised, phantom mask and air mask are used for analysis
                        phantom_mask, air_mask, img_mask = self.Get_Masks(dcm_array)
                        self.mask_dict["phantom"] = phantom_mask
                        self.mask_dict["air"] = air_mask
                        #not suppressed tissue mimic in upper 288 pixels
                        TM_1_mask = phantom_mask.copy()
                        TM_1_mask[288:,:] = False
                        #suppressed tissue mimic in lower 224 pixels
                        TM_2_mask = phantom_mask.copy()
                        TM_2_mask[:288,:] = False
                        self.mask_dict["TM_1"] = TM_1_mask
                        self.mask_dict["TM_2"] = TM_2_mask
                        break
                    break
            except KeyError:
                for series_time in dcm_dict["T1"]:
                    for acq_time in dcm_dict["T1"][series_time]:
                        #image to establish mask from
                        dcm_array = dcm_dict["T1"][series_time][acq_time]["CE"]["pixel_array"].copy()
                        #image mask is the mask visualised, phantom mask and air mask are used for analysis
                        phantom_mask, air_mask, img_mask = self.Get_Masks(dcm_array)
                        self.mask_dict["phantom"] = phantom_mask
                        self.mask_dict["air"] = air_mask
                        #not suppressed tissue mimic in upper 288 pixels
                        TM_1_mask = phantom_mask.copy()
                        TM_1_mask[288:,:] = False
                        #suppressed tissue mimic in lower 224 pixels
                        TM_2_mask = phantom_mask.copy()
                        TM_2_mask[:288,:] = False
                        self.mask_dict["TM_1"] = TM_1_mask
                        self.mask_dict["TM_2"] = TM_2_mask
                        break
                    break
           
        def Fig2Img(self, fig):
            """Convert a Matplotlib figure to a PIL Image and return it"""
            import io
            buf = io.BytesIO()
            fig.savefig(buf)
            buf.seek(0)
            img = Image.open(buf)
            return img
     
            
        def Display_Mask(self, img):
            """
            Parameters
            ----------
            img : PIL Image array
                Image to be displayed in tkinter window
            Return True/ False
            
            Displays img and asks the user to confirm if mask is acceptable.  
            If the initial mask isn't acceptable the thresholds should be 
            adjusted however using different thresholds may indicate acoil failure
            """    
            global img_tk
            top = tk.Toplevel()
            top.title("Is the mask acceptable?")
            top.geometry('800x700')
            top.wm_attributes('-topmost', 1)
            
            #adjust size of image so displays in tkinter label
            scale_factor = 800/img.size[0]
            size = tuple((np.array(img.size) * scale_factor).astype(int))
            img_rs = img.resize(size, resample=Image.NEAREST)
            img_tk = ImageTk.PhotoImage(img_rs)
            tk.Label(top, image=img_tk).grid(row=0, column=0, columnspan = 2)
            proceed = tk.IntVar()
            
            #initialise thresholds
            lower_threshold = tk.DoubleVar()
            upper_threshold = tk.DoubleVar()
            lower_threshold.set(self.lower_threshold)
            upper_threshold.set(self.upper_threshold)
            
            tk.Button(top, text="Yes", command=lambda:  proceed.set(True)).grid(row=1, column=0)
            tk.Button(top, text="Retry", command=lambda:  proceed.set(False)).grid(row=1, column=1)
            tk.Label(top, text = "Lower_threshold").grid(row=2, column=0)
            tk.Entry(top, width = 15, textvariable= lower_threshold).grid(row=2, column=1)
            tk.Label(top, text = "Upper_threshold").grid(row=3, column=0)
            tk.Entry(top, width = 15, textvariable= upper_threshold).grid(row=3, column=1)        
            
            #Wait until the user has decided if the mask is acceptable
            top.wait_variable(proceed)
            top.destroy()
            
            #replace thresholds with those specified by the user
            self.lower_threshold = lower_threshold.get()
            self.upper_threshold = upper_threshold.get()
            return proceed.get()
        
        def Expand_and_Contract_Masks(self, original_mask, x_size, y_size, expand=True):
            """Expand/contract original_mask by 1 pixel in all directions;
            x_size and y_size are the dimensions of the image.
            """
            new_mask = np.full_like(original_mask,False)
            for x in range(x_size):
                for y in range(y_size):
                    if original_mask[x,y] == True:
                        if np.all(original_mask[x-1:x+2,y-1:y+2]):
                            new_mask[x,y] = True
                        else:
                            if expand == True:
                                for i in range(3):
                                    for j in range(3):
                                        new_mask[x-1+i,y-1+j] = True
                            else:
                                new_mask[x,y] = False
            return new_mask
            
        
        def Get_Masks(self,dcm_array):
            """
            

            Parameters
            ----------
            dcm : DICOM
                Combined array DICOM to create mask from

            Returns
            -------
            phantom_mask_npbool : NP Array of Booleans
                img_mask_npbool contracted by 4 pixels. Phantom mask ignoring edge effects   
            air_mask_npbool : NP Array of Booleans
                img_mask_npbool inverted and contracted by 5 pixels. Air mask ignoring edge effects 
            img_mask_npbool : NP Array of Booleans
                Accepted mask covering entirity of phantom
            
            Produces a mask based on previously defined thresholds. 
            Asks the user if the mask is acceptable.  If it isn't the user can 
            adjust the thresholds until the mask is acceptable.  It should be 
            noted if theresholds need adjusting this may indicate a coil failure.
            
            The accepted mask is contracted by 4 pixels to produce a mask of 
            the phantom ignoring edge effects.  The accepted mask is also inverted and 
            econtracted by 5 pixels to produce a mask of the air.
            """
            x_size = dcm_array.shape[0]
            y_size = dcm_array.shape[1]
            max_int = np.amax(dcm_array)
            mask_acceptable = False
            #smooth out noise in the image
            dcm_filtered = cv2.filter2D(dcm_array,-1,self.low_pass_filter)
            
            #Loop until User accepts the mask
            while mask_acceptable ==False:
                markers = np.zeros(dcm_filtered.shape)
                #probably air
                markers[dcm_filtered<max_int*self.lower_threshold]=1
                #probably phantom
                markers[dcm_filtered>max_int*self.upper_threshold]=2
                
                #perform watershed algorithm based on the thresholded image 
                #and the intensities of the voxels in the smoothed DICOM
                labels = seg.watershed(dcm_filtered, markers)
                img_mask = np.full_like(dcm_filtered, False)
                img_mask[labels==2]=True
                
                test = np.ma.masked_where(img_mask==False, img_mask)
                fig, ax = plt.subplots()
                #create figure of mask overlaying the DICOM PixelArray
                ax.imshow(dcm_filtered, cmap='gray', interpolation='none', alpha=0.4)
                ax.imshow(test, 'spring', interpolation='none', alpha=0.2)
                PIL_img = self.Fig2Img(fig)
                #close plot: will be displayed in tkinter window
                plt.close(fig)
                #Display figure in tkinter window and ask user if the mask is OK 
                if self.Display_Mask(PIL_img) == True:
                    mask_acceptable = True
            
            #contract accepted mask by 4 pixels to produce phantom_mask
            phantom_mask = img_mask
            for contract_px in range(4):
                contracted_mask = self.Expand_and_Contract_Masks(phantom_mask, x_size, y_size, expand=False)
                phantom_mask = contracted_mask
            
            
            #Invert accepted mask and contract by 5 pixels to produce air_mask
            air_mask = np.full_like(dcm_filtered, True)
            air_mask[img_mask ==True]=False
            for contract_px in range(5):
                contracted_mask = self.Expand_and_Contract_Masks(air_mask, x_size, y_size, expand=False)
                air_mask = contracted_mask
            
            phantom_mask_npbool = phantom_mask.astype(np.bool_)
            air_mask_npbool = air_mask.astype(np.bool_)
            img_mask_npbool = img_mask.astype(np.bool_)
            
            
            return phantom_mask_npbool, air_mask_npbool, img_mask_npbool    
    
    class calculate_results:
        """Calculate SNR via 3 methods for individual element and combined images:
            -NEMA: SNR calculated via subtraction method; requires 2 images
            -noise_std: SNR calculated using the assumption that the noise in 
            the image is proportional to the S.D. of the signal in air.
            -noise_av (preferred mathod): SNR calculated using the assumption 
            that the noise in the image is proportional to the mean of the 
            signal in air.
        The uniformity of the combined element image is also calculated via 
        the method outlined in IPEM Report 112 and is refereed to as the 
        integral uniformity method. A low pass filter is applied to the image 
        and the "phantom signal mask" (used for SNR calculation) is applied 
        to the image so only voxels only within the phantom are considered. 
        """
        #low pass filter used for uniformity calculation
        low_pass_filter = np.array([[1,2,1],[2,4,2],[1,2,1]])/16
        #constant of proportionality for noise via "noise_std"/"noise_av" 
        #methods depends on the number of elements the image is produced from
        element_scale_factors = {1:{"SD":0.6551,"Mean":1.2533},
                               2:{"SD":0.6824,"Mean":1.8800},
                               3:{"SD":0.6911,"Mean":2.3500},
                               4:{"SD":0.6953,"Mean":2.7416},
                               6:{"SD":0.6994,"Mean":3.3928},
                               8:{"SD":0.7014,"Mean":3.9380},
                               12:{"SD":0.7035,"Mean":4.8271},
                               14:{"SD":0.7040,"Mean":5.2243},
                               15:{"SD":0.7042,"Mean":5.4123},
                               16:{"SD":0.7043,"Mean":5.6128},
                               18:{"SD":0.7046,"Mean":5.9585},
                               20:{"SD":0.7049,"Mean":6.2698},
                               28:{"SD":0.7055,"Mean":7.4425},
                               30:{"SD":0.7056,"Mean":7.7083},
                               32:{"SD":0.7057,"Mean":7.9688},
                               64:{"SD":0.7064,"Mean":11.2916}
                               }
        def __init__(self, dcm_dict, mask_dict, n_elements):
            for image_type in dcm_dict:
                if image_type == "T2":
                    #calculate SNR and uniformity of T2w images
                    self.SNR_results_dict_T2, self.uniformity_results_dict_T2 = self.calculate_SNRs(dcm_dict[image_type], mask_dict, n_elements)
                elif image_type == "T1":
                    #calculate suppression effectiveness and contrast of T1w images  (assumes suppressed images exist)
                    self.sup_results_dict, self.contrast_results_dict = self.calculate_Sup_Effectiveness(dcm_dict["T1"], dcm_dict["T1_sup"], mask_dict)
                    #calculate SNR and uniformity of T1w non suppressed images
                    self.SNR_results_dict_T1, self.uniformity_results_dict_T1 = self.calculate_SNRs(dcm_dict[image_type], mask_dict, n_elements)
                elif image_type == "T1_sup":
                    #calculate SNR and uniformity of T1w suppressed images
                    self.SNR_results_dict_T1_sup, self.uniformity_results_dict_T1_sup = self.calculate_SNRs(dcm_dict[image_type], mask_dict, n_elements)
                
            
            
       
            
        
        
        def SNR_Moriel(self,img_arr,phantom_mask,air_mask,n_elements, bandwidth_scalar):
            """
            Parameters
            ----------
            img_arr : DICOM.PixelArray
                image to calculate SNR of
            phantom_mask : np Boolean array
                mask of phantom ignoring edge effects
            air_mask : np Boolean array
                mask of air ignoring edge effects
            n_elements : int
                Number of elements used to generate img_arr: SNR scale factors 
                depending on number of elements

            Returns
            -------
            SNR_MORIEL_std : Float
                SNR calculated via noise_std method
            SNR_MORIEL_av : Float
                SNR calculated via noise_av method

            Calulates SNR via two methods using a simgle image:
                -noise_std: SNR calculated using the assumption that the noise in 
                the image is proportional to the S.D. of the signal in air.
                -noise_av (preferred mathod): SNR calculated using the assumption 
                that the noise in the image is proportional to the mean of the 
                signal in air.
            """
            #mask img_arr: only phantom
            signal_arr = img_arr[phantom_mask]
            #mask img_arr: only air
            noise_arr = img_arr[air_mask]
            #get scale factors for SNR calculations based on the number of elements
            scale_factors = self.element_scale_factors[n_elements]
            
            signal_av = np.mean(signal_arr)
            
            #scaled noise
            noise_std = np.std(noise_arr)/scale_factors["SD"]
            noise_av = np.mean(noise_arr)/scale_factors["Mean"]
            #calculate SNR
            SNR_MORIEL_std = round(bandwidth_scalar*signal_av/noise_std,2)
            SNR_MORIEL_av = round(bandwidth_scalar*signal_av/noise_av,2)
            return SNR_MORIEL_std, SNR_MORIEL_av
            
        def SNR_NEMA(self,img_arr_1,img_arr_2,phantom_mask, bandwidth_scalar):
            """
            Calculates SNR via NEMA "subtraction" method: requires two 
            identically acquired images (img_arr_1 and img_arr_2) and a 
            mask of the phantom

            """
            signal_arr= img_arr_1[phantom_mask]
            noise_arr = np.int16(np.subtract(np.int32(img_arr_1), np.int32(img_arr_2)))[phantom_mask]
            signal_av = np.mean(signal_arr)
            noise_std = np.std(noise_arr)
            SNR_NEMA = round(bandwidth_scalar*(2**0.5)*signal_av/noise_std,2)
            return SNR_NEMA
             
        def SNR_Calculate_dcm(self, dcm1, phantom_mask, air_mask, dcm2=None, n_elements=1):
            """Calulates SNR via three methods using dcm1 and dmc2.  One 
            result is produced via the NEMA method which requires 2 images and 
            two results are produced via the the noise_std and noise_av: one for
            each DICOM.
            returns SNR_results - a dictionary of the form:
                {"NEMA":<result>, "noise_av":{0:<result>, 1:<result>},
                 "noise_std":{0:<result>, 1:<result>}
                 }
            """
            
            SNR_results = {}
            SNR_results["noise_std"] = {}
            SNR_results["noise_av"] = {}
            #Images should all be acquired at a bandwidth of 222 Hz/Px.  The 
            #scalar allows images which are acqhuired at a different bandwidth to still be used
            try:
                bandwidth_scalar =  ((dcm1.PixelBandwidth/130)**0.5)
            except AttributeError:
                #Some Dicom headers don't possess the atribute PixelBandwidth 
                #(the bandwidth is still in the header but pydicom cant access 
                #via DICOM.PixelBandwidth).  The BW is assumed to be 222 Hz/px
                bandwidth_scalar =  ((222/130)**0.5)
            
            signal_array_1 = dcm1.pixel_array
            SNR_results["noise_std"][0], SNR_results["noise_av"][0] =  self.SNR_Moriel(signal_array_1, phantom_mask, air_mask, n_elements, bandwidth_scalar)
            try:
                signal_array_2 = dcm2.pixel_array
                SNR_results["noise_std"][1], SNR_results["noise_av"][1] =  self.SNR_Moriel(signal_array_2, phantom_mask, air_mask, n_elements, bandwidth_scalar)
                SNR_results["NEMA"] = self.SNR_NEMA(signal_array_1,signal_array_2,phantom_mask, bandwidth_scalar)
            except  IndexError:
                pass
            return SNR_results  
        
        def SNR_Calculate_array(self, signal_array_1, phantom_mask, air_mask, signal_array_2=None, n_elements=1):
            """Calulates SNR via three methods using dcm1 and dmc2.  One 
            result is produced via the NEMA method which requires 2 images and 
            two results are produced via the the noise_std and noise_av: one for
            each DICOM.
            returns SNR_results - a dictionary of the form:
                {"NEMA":<result>, "noise_av":{0:<result>, 1:<result>},
                 "noise_std":{0:<result>, 1:<result>}
                 }
            """
            SNR_results = {}    
            SNR_results["noise_std"] = {}
            SNR_results["noise_av"] = {}
            #The BW is assumed to be 222 Hz/px
            bandwidth_scalar =  ((222/130)**0.5)
            
            SNR_results["noise_std"][0], SNR_results["noise_av"][0] = self.SNR_Moriel(signal_array_1, phantom_mask, air_mask, n_elements, bandwidth_scalar)
            try:
                SNR_results["noise_std"][1], SNR_results["noise_av"][1] =  self.SNR_Moriel(signal_array_2, phantom_mask, air_mask, n_elements, bandwidth_scalar)
                SNR_results["NEMA"] = self.SNR_NEMA(signal_array_1,signal_array_2,phantom_mask, bandwidth_scalar)
            except  IndexError:
                #CHECK CORRECT error
                pass
            return SNR_results
        
        def Uniformity_Calculate(self, dcm_array, phantom_mask):
            """Calculate the uniformity of dcm via 
            the method outlined in IPEM Report 112 and is refereed to as the 
            integral uniformity method. A low pass filter is applied to the image 
            and the "phantom signal mask" (used for SNR calculation) is applied 
            to the image so only voxels only within the phantom are considered. 
            """
            dcm_filtered = cv2.filter2D(dcm_array,-1,self.low_pass_filter)
            ROI_voxels = dcm_filtered[phantom_mask]
            max_int = np.max(ROI_voxels)
            min_int = np.min(ROI_voxels)
            uniformity = 1-((max_int-min_int)/(max_int+min_int))
            return uniformity
        
        def calculate_SNRs(self, dcm_dict_SNR, mask_dict, n_elements):
            """
            Loop through dcm_dict calculating the SNR for all images via 3 methods 
            (NEMA,noise_std and noise_av) and the uniformity of the combined element images
            
            Also calculates group averages for this analysis run  the acquisition ID of these results is "group_av"
            
            each image given a unique acq ID specified by the series time and repeat number
            the NEMA results are placed in the results dictionary with the first repeat acquisition ID
            -------
            Returns
            SNR_results_dict: of the form {snr_type:{<acq. ID>:{"CE":SNR_Results, "IE":{element_n:SNR_Results}}}}}
            uniformity_results_dict: of the form {<acq. ID>:<uniformity>}            
            """
        
            #SNR calculated via 3 methods
            snr_types = ["NEMA", "noise_std", "noise_av"]
            SNR_results_dict = {}
            uniformity_results_dict = {}
            for snr_type in snr_types:
                SNR_results_dict[snr_type] = {}
            
            for series_time in dcm_dict_SNR: 
                
                repeat_numbers = list(dcm_dict_SNR[series_time].keys())
                acq_ID_1 = str(series_time) + "_" + str(repeat_numbers[0])
                acq_ID_2 = str(series_time) + "_" + str(repeat_numbers[1])
                for snr_type in snr_types:
                    SNR_results_dict[snr_type][acq_ID_1] = {}
                    if snr_type != "NEMA":
                        SNR_results_dict[snr_type][acq_ID_2] = {}      
                
                dcm_array_1 = dcm_dict_SNR[series_time][repeat_numbers[0]]["CE"]["pixel_array"]
                dcm_array_2 = dcm_dict_SNR[series_time][repeat_numbers[1]]["CE"]["pixel_array"]
                SNRs = self.SNR_Calculate_array(dcm_array_1, mask_dict["phantom"], mask_dict["air"], dcm_array_2, n_elements)
                
                SNR_results_dict["NEMA"][acq_ID_1]["CE"] = SNRs["NEMA"]
                SNR_results_dict["noise_std"][acq_ID_1]["CE"] = SNRs["noise_std"][0]
                SNR_results_dict["noise_std"][acq_ID_2]["CE"] = SNRs["noise_std"][1]
                SNR_results_dict["noise_av"][acq_ID_1]["CE"] = SNRs["noise_av"][0]
                SNR_results_dict["noise_av"][acq_ID_2]["CE"] = SNRs["noise_av"][1]
                
                uniformity_1 = round(self.Uniformity_Calculate(dcm_array_1, mask_dict["phantom"]),2)
                uniformity_2 = round(self.Uniformity_Calculate(dcm_array_2, mask_dict["phantom"]),2)
                uniformity_results_dict[acq_ID_1] = uniformity_1
                uniformity_results_dict[acq_ID_2] = uniformity_2
                
                try:
                    for element_n in dcm_dict_SNR[series_time][repeat_numbers[0]]["IE"]:
                        dcm_1 = dcm_dict_SNR[series_time][repeat_numbers[0]]["IE"][element_n]["dcm"]
                        dcm_2 = dcm_dict_SNR[series_time][repeat_numbers[1]]["IE"][element_n]["dcm"]
                        SNRs = self.SNR_Calculate_dcm(dcm_1, mask_dict["phantom"], mask_dict["air"], dcm_2, 1)
                        element_ID = "IE_" + str(element_n)
                        SNR_results_dict["NEMA"][acq_ID_1][element_ID] = SNRs["NEMA"]
                        SNR_results_dict["noise_std"][acq_ID_1][element_ID] = SNRs["noise_std"][0]
                        SNR_results_dict["noise_std"][acq_ID_2][element_ID] = SNRs["noise_std"][1]
                        SNR_results_dict["noise_av"][acq_ID_1][element_ID] = SNRs["noise_av"][0]
                        SNR_results_dict["noise_av"][acq_ID_2][element_ID] = SNRs["noise_av"][1]
                except KeyError:
                    #only combined
                    pass
                    
                    
            for result_type in SNR_results_dict:
                acq_IDs = list(SNR_results_dict[result_type])
                SNR_results_dict[result_type]["group_av"] = {}
                
                results = []
                for coil_ID in SNR_results_dict[result_type][acq_IDs[0]]:
                    results = []
                    for acq_ID in acq_IDs:
                        results.append(SNR_results_dict[result_type][acq_ID][coil_ID])
                    SNR_results_dict[result_type]["group_av"][coil_ID] = round(sum(results)/len(results),2)
                
                
            
            results = []
            for acq_ID in uniformity_results_dict:
                results.append(uniformity_results_dict[acq_ID])
            uniformity_results_dict["group_av"] = round(sum(results)/len(results),2)
            
            return SNR_results_dict, uniformity_results_dict
            
            
        def Calculate_TM_Signals(self, dcm_pixel_array, TM1_mask, TM2_mask, phantom_mask, air_mask=None):
            """
            Parameters
            ----------
            dcm_pixel_array : TYPE
                DESCRIPTION.
            TM1_mask : mask of just the non suppressed Tissue mimick
            TM2_mask : mask of just the suppressed Tissue mimick
            phantom_mask : mask covering entire phantom
            air_mask : mask of pixels definitely in air

            Returns
            -------
            TM_averages : {<mask type>: <average masked pixel value in dcm_pixel_array>}

            """
            TM_averages = {}
            TM1_array = dcm_pixel_array[TM1_mask]
            TM2_array = dcm_pixel_array[TM2_mask]
            TM_averages["TM1"] = np.mean(TM1_array)
            TM_averages["TM2"] = np.mean(TM2_array)
            
            phantom_array = dcm_pixel_array[phantom_mask]
            TM_averages["phantom"] = np.mean(phantom_array)
            
            if air_mask.all() == None:
                TM_averages["air"] = None
            else:
                air_array = dcm_pixel_array[air_mask]
                TM_averages["air"] = np.mean(air_array)
            
            return TM_averages
        
        def calculate_Sup_Effectiveness(self, dcm_dict_T1, dcm_dict_T1_sup, mask_dict):
            """
            Calculate suppression effectiveness and contrast of T1w images.
            unique result ID based on series time and repeat number
            
            group averages are also calculated
            
            3 methods of calculating suppression effectiveness:
                 - "Ratio of contrast ratios"
                 - "Ratio of contrasts"
                 - "Scaled Ratio of contrasts"
            Parameters
            ----------
            dcm_dict_T1 : dict of non suppressed pixel arrays
            dcm_dict_T1_sup : dict of non suppressed pixel arrays
            mask_dict : dictionary of masks including phantom, TM1 and TM2 masks

            Returns
            -------
            sup_results_dict : dictionary of suppression effectiveness results
            of the form: {result_ID:{<result type>:<result>}}
            contrast_results_dict : TYPE
                DESCRIPTION.

            """
            
            result_types = ["Suppression Effectiveness", "T1_contrast", "T1_sup_contrast", "TM1_sig", "TM2_sig", "TM2_sup_sig", "TM2_sup_sig"]
            sup_results_dict = {}
            contrast_results_dict = {}
            
            series_times_T1 = list(dcm_dict_T1.keys())
            series_times_T1_sup = list(dcm_dict_T1_sup.keys())
            
            for i in range(len(series_times_T1)):
                
                for k in range(2):
                    supp_eff = {}
                    contrast = {}
                    result_ID = str(series_times_T1[i]) + "_" + str(k)
                    T1_pixel_array = dcm_dict_T1[series_times_T1[i]][k]["CE"]["pixel_array"]
                    T1_sup_pixel_array = dcm_dict_T1_sup[series_times_T1_sup[i]][k]["CE"]["pixel_array"]
                    
                    #calculate average signal in suppressed and non suppressed tissue mimick
                    T1_averages = self.Calculate_TM_Signals(T1_pixel_array, mask_dict["TM_1"], mask_dict["TM_2"], mask_dict["phantom"], mask_dict["air"])
                    T1_sup_averages = self.Calculate_TM_Signals(T1_sup_pixel_array, mask_dict["TM_1"], mask_dict["TM_2"], mask_dict["phantom"], mask_dict["air"])
                    
                    #ratio of average intensities in different tissue mimicks
                    T1_ratio = T1_averages["TM1"]/T1_averages["TM2"]
                    T1_sup_ratio = T1_sup_averages["TM1"]/T1_sup_averages["TM2"]
                    #contrast between different tissue mimicks
                    contrast["Non suppressed"] = T1_averages["TM1"] - T1_averages["TM2"]
                    contrast["Suppressed"] = T1_sup_averages["TM1"] - T1_sup_averages["TM2"]
                    #scale contrast by average pixel intensity within the phantom
                    T1_contrast_scaled = contrast["Non suppressed"]/(T1_averages["phantom"])
                    T1_sup_contrast_scaled = contrast["Suppressed"]/(T1_sup_averages["phantom"])
                    
                    #three methods of calculating suppression effectiveness
                    supp_eff["Ratio of contrast ratios"] = T1_sup_ratio/T1_ratio
                    supp_eff["Ratio of contrasts"] = contrast["Suppressed"]/contrast["Non suppressed"]
                    supp_eff["Scaled ratio of contrasts"] = T1_sup_contrast_scaled/T1_contrast_scaled
                                      
                    contrast_results_dict[result_ID] = contrast
                    sup_results_dict[result_ID] = supp_eff
                    
                    
                    
                    
            #calculate group averages
            result_IDs = list(sup_results_dict.keys())
            sup_results_dict["group_av"] = {}
            contrast_results_dict["group_av"] = {}
            for result_type in sup_results_dict[result_IDs[0]]:
                results = []
                for result_ID in result_IDs:
                    results.append(sup_results_dict[result_ID][result_type])
                sup_results_dict["group_av"][result_type] = round(sum(results)/len(results),2)    
            for result_type in contrast_results_dict[result_IDs[0]]:
                results = []
                for result_ID in result_IDs:
                    results.append(contrast_results_dict[result_ID][result_type])
                contrast_results_dict["group_av"][result_type] = round(sum(results)/len(results),2)    
                
            
            return sup_results_dict, contrast_results_dict
               
    class export_to_excel:
        
        class AlreadyAnalysedError(Exception):
            def __init__(self, message):
                print(message)
            
            pass
        
        def __init__(self, export_path, results, acquisition_date, ask_overwrite_analysis):
            """
            Input SNR, Uniformity, contrast and suppression effectiveness results into 
            a spreadsheet specified by export_path

            Parameters
            ----------
            export_path : path of spreadsheet to export results to
            results : dictionary of results
            acquisition_date : acquisition date of images (assumption made that all images acquired on the same date)
            ask_overwrite_analysis : Bool - if True and results for the images 
            in question have already been inputted the user will be asked of 
            they want to overwrite the analysis.  If False the results are 
            not inputted.

            Returns
            -------
            None.

            """
            if ask_overwrite_analysis.lower() in ["true", "yes", "1"]:
                    self.ask_overwrite_analysis = True
            else:
                self.ask_overwrite_analysis = False
            #headings in all sheets of spreadsheet     
            base_headings = ["acq_date", "acq_time"]
            SNR_results = {}
            uniformity_results = {}
            suppression_results = {}
            
            try:
                SNR_results["T2"] = results.SNR_results_dict_T2
                uniformity_results["T2"] = results.uniformity_results_dict_T2
            except AttributeError:
                #No T2 SNR analysis
                pass
            try:
                SNR_results["T1"] = results.SNR_results_dict_T1
                uniformity_results["T1"] = results.uniformity_results_dict_T1
            except AttributeError:
                #No T1 SNR analysis
                pass  
            try:
                SNR_results["T1_sup"] = results.SNR_results_dict_T1_sup
                uniformity_results["T1_sup"] = results.uniformity_results_dict_T1_sup
            except AttributeError:
                #No T1_sup SNR analysis
                pass     
                
            try:
                suppression_results = results.sup_results_dict
                contrast_results = results.contrast_results_dict
            except AttributeError:
                #No T1 suppression effectiveness analysis
                pass
            
            export_directory = os.path.split(export_path)[0]
            spreadsheet_file_name = os.path.split(export_path)[1]
            self.Open_Spreadsheet(export_directory, spreadsheet_file_name)
            
            #input SNR and uniformity results into spreadsheet
            if SNR_results != {}:
                for image_type in SNR_results:
                    for SNR_result_type in SNR_results[image_type]:
                        result_headings = list(SNR_results[image_type][SNR_result_type]["group_av"].keys())
                        #headings for first row in the sheet
                        sheet_headings = base_headings + result_headings
                        
                        #seperate results sheets for group averaged, need to seperate the data 
                        data_to_export = SNR_results[image_type][SNR_result_type].copy()
                        data_to_export_GA = {}
                        data_to_export_GA[list(data_to_export.keys())[0]] = data_to_export["group_av"]
                        del data_to_export["group_av"]
                        
                        sheet_name =  "SNR_" + image_type + "_" + SNR_result_type
                        self.Open_Sheet(spreadsheet_file_name, sheet_name, sheet_headings)
                        self.Export_Data(data_to_export, acquisition_date, sheet_headings)
                        
                        sheet_name =  "GA_" + sheet_name
                        self.Open_Sheet(spreadsheet_file_name, sheet_name, sheet_headings)
                        self.Export_Data(data_to_export_GA, acquisition_date, result_headings)
                        
                        self.wb.save(spreadsheet_file_name)
                    
                    
                    #headings for first row in the sheet
                    uniformity_result_headings = ["Uniformity"]
                    uniformity_sheet_headings = base_headings + uniformity_result_headings
                    
                    #seperate results sheets for group averaged, need to seperate the data 
                    data_to_export = uniformity_results[image_type].copy()
                    data_to_export_GA = {}
                    data_to_export_GA[list(data_to_export.keys())[0]] = data_to_export["group_av"]
                    del data_to_export["group_av"]
                                                 
                    sheet_name =  "Uniformity_" + image_type 
                    self.Open_Sheet(spreadsheet_file_name, sheet_name, uniformity_sheet_headings)
                    self.Export_Data(data_to_export, acquisition_date, uniformity_result_headings)
                    
                    sheet_name =  "GA_" + sheet_name
                    self.Open_Sheet(spreadsheet_file_name, sheet_name, uniformity_sheet_headings)
                    self.Export_Data(data_to_export_GA, acquisition_date, uniformity_result_headings)
                    
                    self.wb.save(spreadsheet_file_name)
            
            #input contrast and suppression results into spreadsheet
            if suppression_results != {}: 
                data_to_export = suppression_results.copy()
                suppression_results_headings = list(data_to_export["group_av"].keys())
                suppression_sheet_headings = base_headings + suppression_results_headings
                data_to_export_GA = {}
                data_to_export_GA[list(data_to_export.keys())[0]] = data_to_export["group_av"]
                del data_to_export["group_av"]
                 
                
                sheet_name =  "Suppression_T1w" 
                self.Open_Sheet(spreadsheet_file_name, sheet_name, suppression_sheet_headings)
                self.Export_Data(data_to_export, acquisition_date, suppression_results_headings)
                
                sheet_name =  "GA_" + sheet_name
                self.Open_Sheet(spreadsheet_file_name, sheet_name, suppression_sheet_headings)
                self.Export_Data(data_to_export_GA, acquisition_date, suppression_results_headings)
                
                self.wb.save(spreadsheet_file_name)
                
                data_to_export = contrast_results.copy()
                contrast_results_headings = list(data_to_export["group_av"].keys())
                contrast_sheet_headings = base_headings + contrast_results_headings
                data_to_export_GA = {}
                data_to_export_GA[list(data_to_export.keys())[0]] = data_to_export["group_av"]
                del data_to_export["group_av"]
                 
                
                sheet_name =  "Contrast_T1w" 
                self.Open_Sheet(spreadsheet_file_name, sheet_name, contrast_sheet_headings)
                self.Export_Data(data_to_export, acquisition_date, contrast_results_headings)
                
                sheet_name =  "GA_" + sheet_name
                self.Open_Sheet(spreadsheet_file_name, sheet_name, contrast_sheet_headings)
                self.Export_Data(data_to_export_GA, acquisition_date, contrast_results_headings)
                
                
            
           
            self.wb.save(spreadsheet_file_name)
            self.wb.close()
            self.wb.close()
        
        
        
        def Open_Spreadsheet(self, directory, file_name):
            
            """
            Parameters
            ----------
            directory : str
                directory the spreadsheet is contained in
            file_name : str
                name of the saved file
            
            Initialises <self.wb> a workbook (requires openpyxl). If <file_name> 
            doesn't exist in the directory specified it is created.
            """
            
            if not os.path.exists(directory):
                    os.makedirs(directory)
            os.chdir(directory)
            
            """Open correct spreadsheet. If it doesn't exist create it"""
            if os.path.isfile(os.path.join(directory, file_name)):
                self.wb = openpyxl.load_workbook(file_name)
            else:
                #first SNR data for this scanner; create new file
                self.wb = openpyxl.Workbook()
                self.wb.save(file_name)
                print("New Spreadsheet created.  This should be the first SNR measurement for this scanner")
            
    
        def Initialise_Sheet(self, ws_name, headings):
            """
            Parameters
            ----------
            ws_name : str
                Name of worksheet required for initialisation
            headings: list
            
            Create a work sheet in the workbook <self.wb>, called <ws_name>.  
            The works sheet will be initialised with headings specified by 
            <headings>. Initialised <self.ws> (requires openpyxl).
            """
            self.wb.create_sheet(ws_name)
            self.ws = self.wb[ws_name]
            for i in range(len(headings)):
                #write <headings> in the first row of the worksheet
                self.ws.cell(column=i+1, row=1, value=headings[i])
    
        def Open_Sheet(self, file_name, sheet_name, headings_list= None):
            """
            Parameters
            ----------
            file_name : str
            sheet_name : str
            headings_list : list
            
            Initialise self.ws.  If <sheet_name> doesn't exist in <file_name>, 
            create it with headings specified by <headings_list>.
            
            """
            if not sheet_name in self.wb.sheetnames:
                #first SNR data for this coil; create new file
                if headings_list == None:
                    print("Please specify headings_list")
                else:
                    self.Initialise_Sheet(ws_name = sheet_name, headings = headings_list)
                    self.wb.save(file_name)
                    print("New Sheet created.  This should be the first SNR measurement for this coil")
            else:
                self.ws = self.wb[sheet_name]
    
        def Get_Previous_Analysises_Performed(self):
            """
            Returns
            ---------
            previous_results_details : Dictionary of the form{<date>:[<times>]}
                Containing dates and times of previous analysises in self.ws
            headings : list
                List of headings in self.ws
            
            Looks in the first two columns of <self.ws> and returns two lists of all 
            the elements in these to columns.  These columns contain the acquisition 
            dates and acquisition times of the images on which SNR analysis has 
            previously been performed. 
            
            The first row is assumed to contain the headings
            """
            
            dates = {}
            acq_IDs = {} 
            headings = {}
            """loop through rows, column A containes a list of acquisition dates and 
            column B contains a list of acquisiton times"""
            
            for cell in self.ws['A']:
                dates[cell.row] = cell.value
            for cell in self.ws['B']:
                acq_IDs[cell.row] = cell.value
            for cell in self.ws[1]:
                headings[cell.column] = cell.value
              

            return dates, acq_IDs, headings
    
    
        def AskOverwriteAnalysis(self, analysis_details):
            """
            Asks the user if they want to overwrite the analysis
            Parameters
            ----------
            analysis_details : string with details of the analysis to be put in a message to the user
            Returns True/False
            """
            overwrite_analysis = tk.IntVar()
            
            top = tk.Toplevel()
            top.title("Do you want to overwrite the previous analysis?")
            #top.geometry('300x150')
            top.wm_attributes('-topmost', 1)
            tk.Label(top, text=analysis_details,font=("Courier", 16)).grid(columnspan=2, row=0, column=0)
            tk.Label(top, text="Do you want to overwrite the previous analysis?",font=("Courier", 25)).grid(columnspan=2, row=1, column=0)
            tk.Button(top, text="Yes", font=("Courier",16), command=lambda:  overwrite_analysis.set(True)).grid(row=2, column=0)
            tk.Button(top, text="No", font=("Courier",16), command=lambda:  overwrite_analysis.set(False)).grid(row=2, column=1)
            top.wait_variable(overwrite_analysis)          
            top.destroy()
            #Return True/ False
            return overwrite_analysis.get()
    
        def Export_Data(self, results, date, result_headings):   
            """
            Export results to self.ws  raises self.AlreadyAnalysedError if the 
            results have previously been inputted for the date, acq_ID and result type in question
            
            Parameters
            ----------
            results : dictionary of the form{acq_ID:{<result type>:<result>}}
            date: date of image acquisition
            result_headings : list of the types of results - these should eb the headings in the first row of self.ws

            """
            
            #get dates, acq_IDs and headings already in self.ws
            previous_dates, previous_acq_IDs, previous_headings = self.Get_Previous_Analysises_Performed()
            
            previous_headings_col_IDs = list(previous_headings.keys()) 
            #if result type doesn't exist create a new heading for it
            for heading in result_headings:
                if heading not in previous_headings.values():
                    new_col_ID = max(previous_headings_col_IDs) + 1
                    self.ws.cell(column=new_col_ID, row=1, value=heading)
                    previous_headings_col_IDs.append(new_col_ID)
                    
            #loop through results inputting into self.ws with a new row for each acq_ID
            for acq_ID in results:
                try:
                    #input results if not already inputted
                    for row_idx in previous_dates:
                        if previous_dates[row_idx] == date and previous_acq_IDs[row_idx] == acq_ID:
                            error = "Already inputted - WS:" + str(self.ws.title) + " - Date:" + str(date) + " - Acq_ID:" + str(acq_ID)
                            #inform user results not inputted
                            raise self.AlreadyAnalysedError(error)
                            
                    results_to_input = results[acq_ID]
                    row_ID = self.ws.max_row+1
                    self.ws.cell(column=1, row=row_ID, value=date)
                    self.ws.cell(column=2, row=row_ID, value=acq_ID)
                    try:
                        for heading in results_to_input:
                            #multiple headings, please NOTE won't work if duplicate headings
                            heading_col_ID = list(previous_headings.keys())[list(previous_headings.values()).index(heading)] 
                            self.ws.cell(column=heading_col_ID, row=row_ID, value=results_to_input[heading])
                    except TypeError:
                        #single value e.g. uniformity result
                        heading = result_headings[0]
                        heading_col_ID = list(previous_headings.keys())[list(previous_headings.values()).index(heading)] 
                        self.ws.cell(column=heading_col_ID, row=row_ID, value=results_to_input)
                except self.AlreadyAnalysedError:
                    if self.ask_overwrite_analysis == True:
                       overwrite_analysis = self.AskOverwriteAnalysis(error) 
                       if overwrite_analysis == True:
                           results_to_input = results[acq_ID]
                           row_ID = row_idx
                           self.ws.cell(column=1, row=row_ID, value=date)
                           self.ws.cell(column=2, row=row_ID, value=acq_ID)
                           try:
                               for heading in results_to_input:
                                   #multiple headings, please NOTE won't work if duplicate headings
                                   heading_col_ID = list(previous_headings.keys())[list(previous_headings.values()).index(heading)] 
                                   self.ws.cell(column=heading_col_ID, row=row_ID, value=results_to_input[heading])
                           except TypeError:
                               #single value e.g. uniformity result
                               heading = result_headings[0]
                               heading_col_ID = list(previous_headings.keys())[list(previous_headings.values()).index(heading)] 
                               self.ws.cell(column=heading_col_ID, row=row_ID, value=results_to_input)
                           
    class produce_report:
        
        def __init__(self, report_directory, baseline_ss_path, results_ss_path, dcm_dict, results, date, sheet_names, pdf_header_path, pdf_txt_sizes, pdf_default_comment, pdf_glossary, report_sections):
            """
            Produce pdf report summarising results.
            An attempt is made to input SNR, uniformity, supression 
            effectiveness and contrast results.  The appendix contains magnitude 
            images for each element on the day of testing and at commissioning.
            
            An initial report is produced and the user is prompted to input a 
            comment and result (Pass/Fail/Undetermined).  A final report is 
            produced with the comment and result
            
            Parameters
            ----------
            report_directory : path of folder to save report in
            baseline_ss_path : path to spreadsheet of baseline results
            results_ss_path : path to spreadsheet with results
            dcm_dict : dictionary of dicom images containing indicvidual element T2w images
            results : results dictionary from this analysis run
            date : date images acquired for this analysis run (assumption just one date)
            sheet_names : list of sheet names to report (specified in config file) 
            pdf_header_path : path to png to be inserted as the header (specified in config file) 
            pdf_txt_sizes : dictionary of text sizes (specified in config file)
            pdf_default_comment : default comment to be input (specified in config file) 
            pdf_glossary : dictionary of form {<term>: <definition>} (specified in config file) 
            report_sections : dictionary of the form {<section name>: <section description>}
            """
            self.title_size = int(pdf_txt_sizes["title_size"])
            self.heading_size = int(pdf_txt_sizes["heading_size"])
            self.body_txt_size = int(pdf_txt_sizes["body_txt_size"])
            self.default_comment = pdf_default_comment["comment"]
            self.glossary = pdf_glossary
            self.report_sections = report_sections
            self.pdf_header_path = pdf_header_path
            
            self.sheet_names = sheet_names
            
            magnitude_png_directory = os.path.join(report_directory,"magnitude_images_T2",str(date))
            graph_base_png_directory = os.path.join(report_directory,"graphs")
            
            #attempt to produce magnitude images for individual element T2w images
            try:
                T2_series_times = list(dcm_dict["T2"].keys())
                self.Produce_Magnitude_PNGs(magnitude_png_directory, dcm_dict["T2"][T2_series_times[0]][0])
                mag_images_T2 = True
            except KeyError:
                #no T2w magnitude images
                mag_images_T2 = False
                pass
            
            #attempt to produce graphs and results tables for SNR results            
            try:
                SNR_results_T2 = results.SNR_results_dict_T2["noise_av"]["group_av"]
                table_SNR_T2 = self.Produce_Results_Table(SNR_results_T2, baseline_ss_path, self.sheet_names["SNR"])
                self.Create_PNG_Graphs(graph_base_png_directory,baseline_ss_path,results_ss_path, self.sheet_names["SNR"])
            except AttributeError:
                #no T2w  SNR analysis
                table_SNR_T2 = None
            
            #attempt to produce graphs and results tables for uniformity results    
            try:    
                uniformity_results_T2 = {"Uniformity":results.uniformity_results_dict_T2["group_av"]}
                table_uniformity_T2 = self.Produce_Results_Table(uniformity_results_T2, baseline_ss_path, self.sheet_names["uniformity"])
                self.Create_PNG_Graphs(graph_base_png_directory,baseline_ss_path,results_ss_path,  self.sheet_names["uniformity"])
            except AttributeError: 
                #no T2w uniformity analysis
                table_uniformity_T2 = None
           
            #attempt to produce graphs and results tables for suppression effectiveness results 
            try:
                suppression_results = results.sup_results_dict["group_av"]
                table_suppression_T1 = self.Produce_Results_Table(suppression_results, baseline_ss_path, self.sheet_names["suppression"])
                self.Create_PNG_Graphs(graph_base_png_directory,baseline_ss_path,results_ss_path,  self.sheet_names["suppression"])
            except AttributeError:
                #no T1w suppression analysis
                table_suppression_T1 = None      
            
            #attempt to produce graphs and results tables for suppression effectiveness results     
            try:
                contrast_results = results.contrast_results_dict["group_av"]
                table_contrast_T1 = self.Produce_Results_Table(contrast_results, baseline_ss_path, self.sheet_names["contrast"])
                self.Create_PNG_Graphs(graph_base_png_directory,baseline_ss_path,results_ss_path,  self.sheet_names["contrast"])
            except AttributeError:
                #no T1w suppression analysis
                table_contrast_T1 = None   
            
            #create draft pdf  (without comments) for user to review
            self.Create_PDF(report_directory, date, table_SNR_T2, table_uniformity_T2, table_suppression_T1, table_contrast_T1, mag_images_T2)

            #ask user for result and comment
            result_outcome, user_comment = self.Ask_User_Comments()
            
            #create final pdf  (without comments) for user to review
            self.Create_PDF(report_directory, date, table_SNR_T2, table_uniformity_T2, table_suppression_T1, table_contrast_T1, mag_images_T2, result_outcome, user_comment)
           
            
            
        def Initialise_Directory(self, path):
            """
            Checks if the path exists. If it doesn't create it.
            
            Parameters
            ----------
            path : str
                Should be a directory
            """
            if not os.path.exists(path):
                os.makedirs(path)
                
        def PixelArray_to_PNG(self, dcm_array, target_directory, new_name):
            """
            Parameters
            ----------
            dcm_array : np array
                Dicom.pixel_array to produce PNG image from
            target_directory : str
                Directory to save image in
            new_name : str
                Filename to save PNG as   
            Saves the DICOM Pixel array as a PNG
            
            Convert DICOM.PixelArray to a PNG and save it in the specified directory
            """
            #Convert to float to avoid overflow or underflow losses.
            image_2d = dcm_array
        
            # Rescaling grey scale between 0-255
            image_2d_scaled = (np.maximum(image_2d,0) / image_2d.max()) * 255.0
        
            # Convert to uint
            image_2d_scaled = np.uint8(image_2d_scaled)
            
            self.Initialise_Directory(target_directory)
            
            cv2.imwrite(os.path.join(target_directory, new_name), image_2d_scaled)
        
        
        
        def Produce_Magnitude_PNGs(self, magnitude_png_directory, dcm_dict):
            """
            Save magnitude pngs in magnitude_png_directory
            
            Parameters
            ----------
            magnitude_png_directory : str
                directory to save PNGs in 
            dcm_dict : dictionary 
                It is expected that combined element images will be in a pixel 
                array format.  Individual elemet images can be in a pixel array or dicom format
            """
            for image_type in dcm_dict:
                if image_type == "CE":
                    #combined element image  (file name CE.png)
                    dcm_pixel_array = dcm_dict[image_type]["pixel_array"]
                    self.PixelArray_to_PNG(dcm_pixel_array, magnitude_png_directory, "CE" + ".png")
                else:
                    for array_n in dcm_dict[image_type]:
                        #Individual element images
                        try:
                            dcm_pixel_array = dcm_dict[image_type][array_n]["pixel_array"]
                        except KeyError:
                            dcm_pixel_array = dcm_dict[image_type][array_n]["dcm"].pixel_array.astype(float)
                        #name PNGs according to array number
                        self.PixelArray_to_PNG(dcm_pixel_array, magnitude_png_directory, "IE_"+str(array_n) + ".png")
                        
                        
        def Produce_Results_Table(self, results, baseline_ss_path, sheet_name):
            """
            Create table comparing results to previous baselines  Each parameter has a new row in the table with a result and baseline

            Parameters
            ----------
            results : dictionary of the form {<parameter>:<result>}
            baseline_ss_path : path to the spreadsheet containing baseline results
            sheet_name : name of sheet in baseline spreadsheet containing the baseline results

            Returns
            -------
            results_to_print : List whereby each element in results_to_print represents a new row in the table.  
            Each element in results_to_print is a 3 element list.

            """
            #initialise first row of table
            results_to_print = [["Parameter", "Measured Result", "Baseline Result"],]
            try:
                self.wb = openpyxl.load_workbook(baseline_ss_path)
                self.ws = self.wb[sheet_name]
                baseline_ss_headings = {}
                for cell in self.ws[1]:
                    baseline_ss_headings[cell.column] = cell.value
                for parameter in results:
                    #each loop adds a new row to the table
                    measured_value = results[parameter]
                    heading_col_ID = list(baseline_ss_headings.keys())[list(baseline_ss_headings.values()).index(parameter)] 
                    baseline_value = round(self.ws.cell(column=heading_col_ID, row=2).value,2)
                    results_to_print.append([str(parameter), str(measured_value), str(baseline_value)])      
                self.wb.close()
            except FileNotFoundError:
                #no baseline spreadsheet
                for parameter in results:
                    #each loop adds a new row to the table
                    measured_value = results[parameter]
                    baseline_value = "N/A"
                    results_to_print.append([str(parameter), str(measured_value), str(baseline_value)])
            return results_to_print
        
        def Create_PNG_Graphs(self,export_base_dir,baseline_ss_path,results_ss_path,sheet_name):
            """
            create graph of current and historic data. Also plots the baseline values if they exist
            Parameters
            ----------
            export_base_dir : str - directory to save graphs in
            baseline_ss_path : str - path to spreadsheet containing baseline results
            results_ss_path : str - path to spreadsheet containing longitudinal results
            sheet_name : str - sheet name to produce graphs for (same sheet name for both baseline and results spreadsheets)


            """
            
            baseline_dict = {}
            try:
                #Get the baseline results and store in baseline_dict
                self.wb = openpyxl.load_workbook(baseline_ss_path)
                self.ws = self.wb[sheet_name]
                for heading_cell in self.ws[1][2:]:
                    baseline_dict[heading_cell.value] = self.ws.cell(row=2,column=heading_cell.column).value
                self.wb.close()
                baseline_file = True
            except FileNotFoundError:
                baseline_file = False
            
            
            #open results spreadsheet 
            self.wb = openpyxl.load_workbook(results_ss_path)
            self.ws = self.wb[sheet_name]
            #create dictionary so we know what column each heading is in
            results_ss_headings = {}
            for cell in self.ws[1]:
                results_ss_headings[cell.column] = cell.value
            
            
            max_r = self.ws.max_row
            #create dictionary of historical results of the form {<parameter>:[<results>]}
            results = {}
            result_dates = []
            result_dates_formatted = []
            for r in range(2,max_r+1):
                date = self.ws.cell(row=r, column=1).value
                result_dates.append(date)
                #date in dd/mm/yyyy format
                result_dates_formatted.append(date[-2:]+"/"+date[4:-2]+"/"+date[2:-4])
                
            if baseline_file == True:
                for heading in baseline_dict:
                    #heading is the paramater measured
                    heading_col_ID = list(results_ss_headings.keys())[list(results_ss_headings.values()).index(heading)]
                    heading_results = []
                    for r in range(2,max_r+1):
                        heading_results.append(self.ws.cell(row=r, column=heading_col_ID).value)
                    results[heading]  = heading_results  
            else:
                #no baselines established
                for heading_col_ID in results_ss_headings:
                    if heading_col_ID >2:
                        heading = results_ss_headings[heading_col_ID]
                        heading_results = []
                        for r in range(2,max_r+1):
                            heading_results.append(self.ws.cell(row=r, column=heading_col_ID).value)
                        results[heading]  = heading_results
                
            self.wb.close()
            
            #create graph for parameters (headings) 
            for heading in results:
                png_file_name = heading + ".png"
                export_dir = os.path.join(export_base_dir, str(result_dates[-1]), sheet_name)
                #ensure directory exists
                self.Initialise_Directory(export_dir)
                png_file_path = os.path.join(export_dir, png_file_name)
                
                fig, ax = plt.subplots(nrows=1, ncols=1)
                fig.set_tight_layout(True)
                
                
                    
                ax.title.set_text(heading + " " + sheet_name)
                ax.set_ylabel(sheet_name)
                ax.set_xlabel("Date")
                
                x_values = np.arange(max_r-1)
                ax.plot(x_values, results[heading], label='Results')
                try:
                    #add lines showing +/- 10% from baseline
                    baseline_value = baseline_dict[heading]
                    ax.hlines(y=baseline_value*0.9, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle="--", label='Baseline -10%')
                    ax.hlines(y=baseline_value*1.1, xmin=x_values[0], xmax=x_values[-1], linewidth=2, color='r', linestyle=":", label='Baseline +10%')
                except KeyError:
                    #no baseline value
                    pass
                ax.set_xticks(x_values)
                #label with dates of analysis
                ax.set_xticklabels(result_dates_formatted, rotation=45)
                ax.legend()
                fig.savefig(png_file_path)
                plt.close(fig)
         
        def Input_Table_PDF(self,table_data,table_title=None):
            """
            Input table into a pdf
            Parameters
            ----------
            table_data : List whereby each element in results_to_print represents a new row in the table.  
            Each element in results_to_print is a list with each element being in a new column.
            table_title : str - to be put in a merged first row of the table

            Returns
            -------
            None.

            """
            n_cols = len(table_data[0])
            n_rows = len(table_data)
            #If a title is specified the first row full width containing the title 
            if table_title != None:
                self.pdf.set_font('arial', 'B', self.heading_size)
                self.pdf.cell(self.pdf.epw, self.pdf.font_size*2.5, table_title, border=1, ln=1, align='C')
            self.pdf.set_font('arial', 'B', self.body_txt_size)#bold headings
            table_line_height = self.pdf.font_size * 2.5
            table_col_width = self.pdf.epw / n_cols 
            # add new row for each element in table data
            for row in table_data:
                for datum in row:
                    self.pdf.cell(table_col_width, table_line_height, datum, border=1, ln=0, align='C')
                self.pdf.set_font('arial', '', self.body_txt_size)
                self.pdf.ln()  
            
        def Input_Graphs_PDF(self, results_graphs_dir):
            """
            Input graphs in results_graphs_dir into the pdf in two columns
            Parameters
            ----------
            results_graphs_dir : str - directory containing pngsof graphs to be inputted
            """
            graphs_file_paths = {}
            graphs_file_names = [f for f in os.listdir(results_graphs_dir) if os.path.isfile(os.path.join(results_graphs_dir,f))]
            #get file names to label graphs in pdf
            for file_name in graphs_file_names:
                if file_name[-4:] == ".png":
                    graphs_file_paths[file_name] = os.path.join(results_graphs_dir,file_name)
            
            self.pdf.set_font('arial', 'B', self.body_txt_size) 
            #ensure enough room for label and graph
            self.pdf.set_auto_page_break(True, margin=(self.body_txt_size*2.5)+(4*self.pdf.epw/16))
            for i in range(1+(int(len(graphs_file_names)/2))):
                idx_1 = i*2
                idx_2 = i*2 + 1
                if idx_2 <len(graphs_file_names):
                    #if 2 graphs left to input put them side by side
                    self.pdf.set_font('arial', 'B', self.body_txt_size)
                    self.pdf.cell(w=self.pdf.epw/2, h=self.pdf.font_size*2.5, txt=graphs_file_names[idx_1][:-4], border=0,ln=0, align='C')
                    self.pdf.cell(w=self.pdf.epw/2, h=self.pdf.font_size*2.5, txt=graphs_file_names[idx_2][:-4], border=0,ln=1, align='C')
                    current_x = self.pdf.get_x()
                    current_y = self.pdf.get_y()
                    self.pdf.image(graphs_file_paths[graphs_file_names[idx_1]],x=current_x+(self.pdf.epw/16),y=current_y,w=6*self.pdf.epw/16)
                    self.pdf.image(graphs_file_paths[graphs_file_names[idx_2]],x=current_x+(9*self.pdf.epw/16),y=current_y,w=6*self.pdf.epw/16)
                    self.pdf.set_y(current_y+4*self.pdf.epw/16)
                elif idx_1 <len(graphs_file_names):
                    #center final graph if there are an odd number
                    self.pdf.set_font('arial', 'B', self.body_txt_size)
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt=graphs_file_names[idx_1][:-4], border=0,ln=1, align='C')
                    current_x = self.pdf.get_x()
                    current_y = self.pdf.get_y()
                    self.pdf.image(graphs_file_paths[graphs_file_names[idx_1]],x=current_x+(5*self.pdf.epw/16),y=current_y,w=6*self.pdf.epw/16)
                    self.pdf.set_y(current_y+4*self.pdf.epw/16)
            #reset autopage break to default
            self.pdf.set_auto_page_break(True, margin=20)
        
        def PleaseCloseFile(self, pdf_file_path):
            """Called if the a pdf of the same file name is left open when trying to save the new pdf"""
            proceed = tk.IntVar()               
            while proceed.get() != True:
                top = tk.Toplevel()
                top.title("Please Input Outcome and Comments")
                #top.geometry('300x150')
                top.wm_attributes('-topmost', 1)
                tk.Label(top, text="Please close pdf then press continue:",font=("Courier", 25)).grid(columnspan=3, row=0, column=0)
                tk.Label(top, text=pdf_file_path,font=("Courier", 16)).grid(columnspan=3, row=1, column=0)
                tk.Button(top, text="Continue", font=("Courier",16), command=lambda:  proceed.set(True)).grid(row=2, column=1)
                top.wait_variable(proceed)          
                top.destroy()
        
        def Create_PDF(self, report_directory, date, table_SNR_T2, table_uniformity_T2, table_suppression_T1, table_contrast_T1, mag_images=True, result=None, comment=None):
            """
            Produce pdf report of results in report_directory
            Parameters
            ----------
            report_directory : str
            date : str Date of analysis to be inputted in the form yyyymmdd
            
            All tables are represented by a list of lists and  in the same format.
            Each element in the list represents a new row in the table. With each element in the element representing a new column 
            table_SNR_T2 : SNR results 
            table_uniformity_T2 : Uniformity results 
            table_suppression_T1 : Supression Results 
            table_contrast_T1 : Contrast results
            mag_images : Bool - if True input magnitude images into pdf
            result : str - Either "Pass", "Fail" or "Undetermined" - Analysis result
            comment : str - Summary of any comments from the analysis
            Returns
            -------
            pdf_file_path : str

            """
            
            base_graphs_dir = os.path.join(report_directory, "graphs", str(date))
            #date in format ddmmyyyy
            display_date = date[-2:]+"/"+date[4:-2]+"/"+date[2:-4]
            
            #try to initialise pdf with header specified by self.pdf_header_path
            class PDF(FPDF):
                def header(this):
                    # Logo
                    try:
                        this.image(self.pdf_header_path,this.l_margin,this.t_margin,w=this.epw)                    
                        this.ln()
                        this.ln(this.t_margin+self.body_txt_size)
                    except FileNotFoundError:
                        pass
            
            self.pdf = PDF()
            self.pdf.set_xy(0, 0)
            self.pdf.set_font('arial', 'B', self.title_size)
            self.pdf.epw = self.pdf.w-self.pdf.l_margin-self.pdf.r_margin
            self.pdf.eph = self.pdf.h-self.pdf.t_margin-self.pdf.b_margin
            self.pdf.add_page()
            
            #Title of report
            self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="NHSBSP MRI QC Report "+str(display_date), border=0,ln=1, align='C')
            
            #If a result is provided print it
            if result != None:
                self.pdf.cell(w=self.pdf.epw/2, h=self.pdf.font_size*2.5, txt="Result:", border=0,ln=0, align='C')
                if result == "PASS":
                    self.pdf.set_text_color(r=0,g=255,b=0)
                elif result == "FAIL":
                    self.pdf.set_text_color(r=255,g=0,b=0)
                elif result == "Undetermined":
                    self.pdf.set_text_color(r=255,g=128,b=0)
                self.pdf.cell(w=self.pdf.epw/2, h=self.pdf.font_size*2.5, txt=result, border=0,ln=1, align='C')
                self.pdf.set_text_color(0)
            #If a comment is provided print it
            if comment != None:
                self.pdf.set_font('arial', 'B', self.heading_size)
                self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="Comments:", border=1, ln=1, align='C')
                self.pdf.set_font('arial', 'B', self.body_txt_size)
                self.pdf.multi_cell(w=self.pdf.epw, h=self.pdf.font_size*1.5, txt=comment, border=1, align='L')
                
            self.pdf.set_font('arial', 'B', self.heading_size)
            #Print section title
            self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="Report Sections:", border=0,ln=1, align='L')
            #self.report_sections in the form {section title : description}
            for item in self.report_sections:
                self.pdf.set_font('arial', 'B', self.body_txt_size)
                self.pdf.cell(self.pdf.epw, h=self.pdf.font_size*1.5, txt=item, border=0,ln=1, align='L')
                self.pdf.set_font('arial', '', self.body_txt_size)
                self.pdf.multi_cell(w=self.pdf.epw, h=self.pdf.font_size*1.5, txt=self.report_sections[item], border=0, align='L')            
            
            self.pdf.set_font('arial', 'B', self.heading_size)
            self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="Glossary:", border=0,ln=1, align='L')
            #self.glossary in the form {term : definition}
            for item in self.glossary:
                self.pdf.set_font('arial', 'B', self.body_txt_size)
                self.pdf.cell(w=self.pdf.epw/4, h=self.pdf.font_size*1.5, txt=item, border=0,ln=0, align='L')
                self.pdf.set_font('arial', '', self.body_txt_size)
                self.pdf.multi_cell(w=3*self.pdf.epw/4, h=self.pdf.font_size*1.5, txt=self.glossary[item], border=0, align='L')
            
            #try to input SNR and uniformity results and graphs
            try:
                if table_SNR_T2 != None and table_uniformity_T2 != None:
                    self.pdf.add_page()
                    self.pdf.set_font('arial', 'B', self.heading_size)
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="T2W SE Image Analysis", border=0,ln=1, align='L')
                    self.Input_Table_PDF(table_SNR_T2, "SNR Results")
                    SNR_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["SNR"])
                    self.pdf.add_page()
                    self.Input_Graphs_PDF(SNR_graphs_dir)
                    self.pdf.add_page()
                    self.Input_Table_PDF(table_uniformity_T2, "Uniformity Results")
                    uniformity_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["uniformity"])
                    self.Input_Graphs_PDF(uniformity_graphs_dir)                
                elif table_SNR_T2 != None:
                    #only input SNR results
                    self.pdf.add_page()
                    self.pdf.set_font('arial', 'B', self.heading_size)
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="T2W SE Image Analysis", border=0,ln=1, align='L')
                    self.Input_Table_PDF(table_SNR_T2, "SNR Results")
                    SNR_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["SNR"])
                    self.pdf.add_page()
                    self.Input_Graphs_PDF(SNR_graphs_dir)
                elif table_uniformity_T2 != None:
                    #only input Suniformity results results
                    self.pdf.add_page()
                    self.pdf.set_font('arial', 'B', self.heading_size)
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="T2W SE Image Analysis", border=0,ln=1, align='L')
                    self.Input_Table_PDF(table_uniformity_T2, "Uniformity Results")
                    uniformity_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["uniformity"])
                    self.Input_Graphs_PDF(uniformity_graphs_dir)
            except FileNotFoundError:
                #no graphs for specified date, images being analysed are likely historic (graphs are dated with last date in the results spreadsheet).
                pass
             
            #try to input suppression effectiveness and contrast results    
            try:
                if table_suppression_T1 != None and table_contrast_T1 != None:
                    self.pdf.add_page()
                    self.pdf.set_font('arial', 'B', self.heading_size) 
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="T1W GRE Image Analysis", border=0,ln=1, align='L')
                    self.Input_Table_PDF(table_suppression_T1, "Suppression Effectiveness Results")
                    suppression_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["suppression"])
                    self.Input_Graphs_PDF(suppression_graphs_dir)
                    self.pdf.add_page()
                    self.Input_Table_PDF(table_contrast_T1, "Contrast Results")
                    contrast_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["contrast"])
                    self.Input_Graphs_PDF(contrast_graphs_dir)
                elif table_suppression_T1 != None:
                    #only input suppression effectiveness results
                    self.pdf.add_page()
                    self.pdf.set_font('arial', 'B', self.heading_size) 
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="T1W GRE Image Analysis", border=0,ln=1, align='L')
                    self.Input_Table_PDF(table_suppression_T1, "Suppression Effectiveness Results")
                    suppression_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["suppression"])
                    self.Input_Graphs_PDF(suppression_graphs_dir)
                elif table_contrast_T1 != None:
                    #only input suppression contrast results
                    self.pdf.add_page()
                    self.pdf.set_font('arial', 'B', self.heading_size) 
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt="T1W GRE Image Analysis", border=0,ln=1, align='L')
                    self.Input_Table_PDF(table_contrast_T1, "Contrast Results")
                    contrast_graphs_dir = os.path.join(base_graphs_dir, self.sheet_names["contrast"])
                    self.Input_Graphs_PDF(contrast_graphs_dir)
            except FileNotFoundError:
                #no graphs for specified date, images being analysed are likely historic (graphs are dated with last date in the results spreadsheet).
                pass    
            
            if mag_images == True:
                #input magnitude images on date of testing and baseline side by side
                self.pdf.add_page()
                baseline_mag_images_dir = os.path.join(report_directory, "magnitude_images_T2", "Baseline")
                results_mag_images_dir = os.path.join(report_directory, "magnitude_images_T2", str(date))
                #ensure both headings and magnitude images can fit on the page
                self.pdf.set_auto_page_break(True, margin=(self.body_txt_size*2.5)+(6*self.pdf.epw/16))
                self.pdf.set_font('arial', 'B', self.heading_size)
                self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt=" Appendix A - T2W SE Magnitude Images", border=0,ln=1, align='C')
            
                mag_file_paths = {}
                results_mag_images_file_names = [f for f in os.listdir(results_mag_images_dir) if os.path.isfile(os.path.join(results_mag_images_dir,f))]
                for file_name in results_mag_images_file_names:
                    #baseline_magnitude images
                    mag_file_paths[file_name] = {}
                    mag_file_paths[file_name]["Measured"] = os.path.join(results_mag_images_dir,file_name)
                    mag_file_paths[file_name]["Baseline"] = os.path.join(baseline_mag_images_dir,file_name)
                   
                for image_type in mag_file_paths:
                    self.pdf.set_font('arial', 'B', self.body_txt_size)
                    self.pdf.cell(w=self.pdf.epw, h=self.pdf.font_size*2.5, txt=image_type[:-4], border=0,ln=1, align='C')
                    self.pdf.set_font('arial', '', self.body_txt_size)
                    self.pdf.cell(w=self.pdf.epw/2, h=self.pdf.font_size*2.5, txt="Baseline", border=0,ln=0, align='C')
                    self.pdf.cell(w=self.pdf.epw/2, h=self.pdf.font_size*2.5, txt=str(date), border=0,ln=1, align='C')
                    current_x = self.pdf.get_x()
                    current_y = self.pdf.get_y()
                    try:
                        self.pdf.image(mag_file_paths[image_type]["Baseline"],x=current_x+(self.pdf.epw/16),y=current_y,w=6*self.pdf.epw/16)
                    except FileNotFoundError:
                        #no baseline magnitude image
                        pass
                    self.pdf.image(mag_file_paths[image_type]["Measured"],x=current_x+(9*self.pdf.epw/16),y=current_y,w=6*self.pdf.epw/16)
                    self.pdf.set_y(current_y+6*self.pdf.epw/16)
                #reset auto page break to default
                self.pdf.set_auto_page_break(True, margin=20)
            
            #Report can't be produced if a pdf of the same filename is open.  Get the user to close the file if it is open
            draft_report = False
            pdf_file_name = "NHSBSP_MRI_Report_" + str(date)+".pdf"
            pdf_file_path = os.path.join(report_directory, pdf_file_name)
            while draft_report == False:
                try:
                    self.pdf.output(pdf_file_path, 'F')
                    draft_report = True
                except PermissionError:
                    #user still has the draft report open  (needs to be closed before the final report can be produced)
                    self.PleaseCloseFile(pdf_file_path)
            #open preview of report
            webbrowser.open_new(pdf_file_path)
            return pdf_file_path
            
        def Ask_User_Comments(self):
            """Ask the user for the overall result outcome and any comments to be inputted into the pdf"""
            
            result_outcome = None
            user_comment = self.default_comment
            multiple_outcomes = False
            no_outcomes = False
            while result_outcome == None:
                input_comments = tk.IntVar()               
                while input_comments.get() != True:
                    top = tk.Toplevel()
                    top.title("Please Input Outcome and Comments")
                    #top.geometry('300x150')
                    top.wm_attributes('-topmost', 1)
                    
                    #three results outcomes (pass, fail and undetermined)
                    pass_result = tk.IntVar()
                    pass_result.set(False)
                    fail_result = tk.IntVar()
                    fail_result.set(False)
                    undetermined_result = tk.IntVar()
                    undetermined_result.set(False)
                    
                    comments = tk.IntVar()
                    comments.set(self.default_comment)

                    tk.Label(top, text="Outcome",font=("Courier", 25)).grid(columnspan=3, row=0, column=0)
                    tk.Checkbutton(top, text="Pass", variable=pass_result, font=("Courier", 15)).grid(row=1, column=0)
                    tk.Checkbutton(top, text="Fail", variable=fail_result, font=("Courier", 15)).grid(row=1, column=1)
                    tk.Checkbutton(top, text="Undetermined", variable=undetermined_result, font=("Courier", 15)).grid(row=1, column=2)
                    
                    #ensure ONE outcome is selected
                    if multiple_outcomes == True:
                        tk.Label(top, text="Please select only ONE outcome.",font=("Courier", 25)).grid(columnspan=3, row=2, column=0)
                        multiple_outcomes = False
                    elif no_outcomes == True:
                        tk.Label(top, text="Please select an outcome.",font=("Courier", 25)).grid(columnspan=3, row=2, column=0)
                        no_outcomes = False
                    ttk.Separator(top, orient="horizontal").grid(columnspan=3,row=3, sticky="ew")
                    
                    #space for user to edit the default comment
                    tk.Label(top, text="Comments",font=("Courier", 25)).grid(columnspan=3, row=4, column=0)
                    t = tk.Text(top, height = 5, width = 100)
                    t.grid(columnspan=3, row=5, column=0)
                    t.delete(1.0,"end")
                    t.insert(1.0,user_comment)
                    ttk.Separator(top, orient="horizontal").grid(columnspan=3,row=6, sticky="ew")
                              
                    tk.Button(top, text="Input comments", font=("Courier",15), command=lambda:  input_comments.set(True)).grid(row=7, column=1)
                    
                    top.wait_variable(input_comments)                   
                    user_comment = t.get(1.0,"end")
                    top.destroy()
                if pass_result.get() == True:
                    if fail_result.get() == False and undetermined_result.get() == False:
                        result_outcome = "PASS"
                    else:
                        multiple_outcomes = True
                elif fail_result.get() == True:
                    if pass_result.get() == False and undetermined_result.get() == False:
                        result_outcome = "FAIL"
                    else:
                        multiple_outcomes = True
                elif undetermined_result.get() == True:
                    if pass_result.get() == False and fail_result.get() == False:
                        result_outcome = "Undetermined"
                    else:
                        multiple_outcomes = True
                else:
                    no_outcomes = True
                
            
            return result_outcome, user_comment
        
    def __init__(self, master):
        """
        Parameters
        ----------
        master : tkinter root
        """
        master.withdraw()
        #read config file to obtain default parameters for analysis
        config = configparser.ConfigParser()
        config.optionxform = str
        config.read(self.config_file_name)
    
        initialised = self.initialise_analysis(config["default_paths"], config["sequence_names"], int(config["coil_spec"]["Elements"]))
        masks = self.initialise_masks(initialised.sorted_dcm_dict, config["coil_spec"]).mask_dict
        results = self.calculate_results(initialised.sorted_dcm_dict, masks, int(config["coil_spec"]["Elements"]))
        #if required export results to excel
        if initialised.export_results == True:
            self.export_to_excel(initialised.results_path, results, initialised.acquisition_dates[0], config["user_input"]["ask_overwrite_analysis"])
        #if required produce report of results
        if initialised.produce_report == True:
            self.produce_report(initialised.report_path, initialised.baseline_path, initialised.results_path, initialised.sorted_dcm_dict, results, initialised.acquisition_dates[0], config["sheet_names_to_report"], config["pdf_header_image"]["file_path"], config["pdf_text_sizes"], config["pdf_default_comment"], config["pdf_glossary"], config["report_sections"])
        master.destroy()
        
main(root)
root.mainloop()